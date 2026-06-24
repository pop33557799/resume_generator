"""
Resume Generator — produces an ATS-optimized resume tailored to a job description.

Model (per the reference prompt):
  - You keep a FIXED set of facts: name + contact, your companies (name, location,
    start/end dates), and your university (name, location, dates).
  - You paste a job description.
  - The OpenAI API writes everything else, tailored to that job: a clean job-title
    SUBTITLE, a seniority-matched Summary, 4-6 Skills categories, and for each company
    a JD-appropriate TITLE, a one-line DESCRIPTION, and quantified achievement bullets.
    It also picks a DEGREE name that fits the job.
  - Dates, locations, and the school are taken VERBATIM from your profile. Company
    names can come from the profile or from AI-generated mid-sized employers.
  - Output is rendered to the Bitter/Roboto PDF format defined in resume_format.py.
"""

import json
import os
import re
import html
import tempfile
import time
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from dotenv import load_dotenv
from flask import (
    Flask, abort, flash, redirect, render_template, request,
    send_from_directory, session, url_for,
)
from markupsafe import Markup
from openai import OpenAI
from weasyprint import HTML

load_dotenv()

BASE_DIR = Path(__file__).parent
LEGACY_PROFILE_PATH = BASE_DIR / "profile.json"   # single-profile file (auto-migrated)
PROFILES_DIR = BASE_DIR / "profiles"              # one <id>.json per profile
ACTIVE_PROFILE_FILE = BASE_DIR / ".active_profile"
RESUMES_DIR = BASE_DIR / "resumes"
LEGACY_RESUMES_LOG = BASE_DIR / "resumes_log.xlsx"   # pre-multi-profile shared log
LOG_HEADERS = ["Date", "Job Title", "URL", "Company", "Filename", "Progress", "Note"]


def resumes_log_path(pid: str) -> Path:
    """Each profile keeps its own log file: resumes_log_<id>.xlsx."""
    return BASE_DIR / f"resumes_log_{pid}.xlsx"


def _migrate_legacy_log() -> None:
    """One-time: assign the old shared resumes_log.xlsx to the first profile."""
    if not LEGACY_RESUMES_LOG.exists():
        return
    pid = get_active_id()
    if pid and not resumes_log_path(pid).exists():
        os.replace(LEGACY_RESUMES_LOG, resumes_log_path(pid))
MODEL = os.getenv("OPENAI_MODEL", "gpt-4o")

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET", "dev-secret-change-me")

_MONTHS = ["", "January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"]


# ---------------------------------------------------------------------------
# Profile persistence
# ---------------------------------------------------------------------------

def empty_profile() -> dict:
    return {"name": "", "industry": "", "email": "", "phone": "", "location": "",
            "links": [], "companies": [], "education": []}


def _profile_path(pid: str) -> Path:
    return PROFILES_DIR / f"{pid}.json"


def _unique_pid(name: str) -> str:
    base = _slug(name).lower() or "profile"
    pid, n = base, 2
    while _profile_path(pid).exists():
        pid, n = f"{base}_{n}", n + 1
    return pid


def _migrate_legacy_profile() -> None:
    """One-time: move the old single profile.json into profiles/ on first run."""
    PROFILES_DIR.mkdir(exist_ok=True)
    if any(PROFILES_DIR.glob("*.json")):
        return
    if LEGACY_PROFILE_PATH.exists():
        data = json.loads(LEGACY_PROFILE_PATH.read_text(encoding="utf-8"))
        pid = _unique_pid(data.get("name") or "profile")
        save_profile(data, pid)
        set_active_id(pid)
        LEGACY_PROFILE_PATH.rename(BASE_DIR / "profile.migrated.json")


def list_profiles() -> list:
    """[{id, name}] for every saved profile, sorted by name."""
    _migrate_legacy_profile()
    out = []
    for p in PROFILES_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            continue
        industry = (data.get("industry") or "").strip()
        name = data.get("name") or "(unnamed)"
        label = f"{name} - {industry}" if industry else name
        out.append({"id": p.stem, "name": name, "industry": industry, "label": label})
    out.sort(key=lambda x: x["label"].lower())
    return out


def list_profile_names(profiles: list | None = None) -> list:
    """Unique users, each pointing at that user's first saved profile."""
    profiles = profiles or list_profiles()
    names, seen = [], set()
    for p in profiles:
        if p["name"] in seen:
            continue
        seen.add(p["name"])
        names.append({"name": p["name"], "id": p["id"]})
    return names


def profiles_for_name(name: str, profiles: list | None = None) -> list:
    """All saved profiles for one user/name. Kept for old per-industry logs."""
    profiles = profiles or list_profiles()
    return [p for p in profiles if p["name"] == name]


def get_active_id() -> str | None:
    if ACTIVE_PROFILE_FILE.exists():
        pid = ACTIVE_PROFILE_FILE.read_text(encoding="utf-8").strip()
        if pid and _profile_path(pid).exists():
            return pid
    profiles = list_profiles()
    return profiles[0]["id"] if profiles else None


def set_active_id(pid: str | None) -> None:
    ACTIVE_PROFILE_FILE.write_text(pid or "", encoding="utf-8")


def load_profile(pid: str | None = None) -> dict:
    """Load a profile by id; None = the active profile. Empty if none exist."""
    _migrate_legacy_profile()
    if pid is None:
        pid = get_active_id()
    if pid and _profile_path(pid).exists():
        data = json.loads(_profile_path(pid).read_text(encoding="utf-8"))
        return {**empty_profile(), **data}
    return empty_profile()


def save_profile(profile: dict, pid: str) -> None:
    PROFILES_DIR.mkdir(exist_ok=True)
    _profile_path(pid).write_text(
        json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8")


def create_profile(name: str) -> str:
    pid = _unique_pid(name or "profile")
    save_profile({**empty_profile(), "name": name, "industry": ""}, pid)
    return pid


def delete_profile(pid: str) -> None:
    p = _profile_path(pid)
    if p.exists():
        p.unlink()


def parse_profile_form(form) -> dict:
    profile = empty_profile()
    for key in ("name", "email", "phone", "location"):
        profile[key] = form.get(key, "").strip()

    for label, url in zip(form.getlist("link_label"), form.getlist("link_url")):
        if label.strip() or url.strip():
            profile["links"].append({"label": label.strip(), "url": url.strip()})

    names = form.getlist("comp_name")
    locs = form.getlist("comp_location")
    starts = form.getlist("comp_start")
    ends = form.getlist("comp_end")
    n_companies = max(len(names), len(locs), len(starts), len(ends))
    for i in range(n_companies):
        name = names[i].strip() if i < len(names) else ""
        location = locs[i].strip() if i < len(locs) else ""
        start = starts[i].strip() if i < len(starts) else ""
        end = ends[i].strip() if i < len(ends) else ""
        if not (name or location or start or end):
            continue
        profile["companies"].append({
            "name": name or f"Experience {len(profile['companies']) + 1}",
            "location": location,
            "start": start,
            "end": end,
        })

    unis = form.getlist("edu_university")
    elocs = form.getlist("edu_location")
    estarts = form.getlist("edu_start")
    eends = form.getlist("edu_end")
    for i in range(len(unis)):
        if not unis[i].strip():
            continue
        profile["education"].append({
            "university": unis[i].strip(),
            "location": elocs[i].strip() if i < len(elocs) else "",
            "start": estarts[i].strip() if i < len(estarts) else "",
            "end": eends[i].strip() if i < len(eends) else "",
        })
    return profile


# ---------------------------------------------------------------------------
# Date / keyword formatting
# ---------------------------------------------------------------------------

def fmt_date(value: str, is_end: bool = False) -> str:
    """MM/YYYY -> 'Month YYYY'; YYYY stays; blank/Present -> 'Present' for an end."""
    value = (value or "").strip()
    if not value or value.lower() == "present":
        return "Present" if is_end else value
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", value)
    if m:
        month = int(m.group(1))
        if 1 <= month <= 12:
            return f"{_MONTHS[month]} {m.group(2)}"
    return value


def fmt_range(start: str, end: str) -> str:
    s, e = fmt_date(start), fmt_date(end, is_end=True)
    return f"{s} – {e}" if s and e else (s or e)


# The prompt forbids percentages, but the model occasionally emits a delta like
# "by 20%". Strip that clause and tidy the result (keeps endorsed figures such as
# "99.9% uptime", which aren't deltas).
_PERCENT_RE = re.compile(r"\s*\bby\s+\d+(?:\.\d+)?\s*(?:%|percent\b)", re.IGNORECASE)


def strip_percentages(text: str) -> str:
    if not text:
        return text
    cleaned = _PERCENT_RE.sub("", text)
    cleaned = re.sub(r"\s+([.,;])", r"\1", cleaned)   # " ." -> "."
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return cleaned.strip()


def build_warnings(tailored: dict) -> list:
    """Flag likely parse/format problems so the user can regenerate instead of
    silently shipping a sparse resume."""
    warnings = []
    if not tailored["summary"]:
        warnings.append("The Summary came back empty — the model output may be malformed.")
    if not tailored["skills"]:
        warnings.append("No skill categories were parsed.")
    minimums = [10, 8, 6] + [4] * 12
    for i, e in enumerate(tailored["experience"]):
        n = len(e["bullets"])
        if n == 0:
            warnings.append(f"{e['company']}: no bullet points were parsed.")
        elif i < len(minimums) and n < minimums[i]:
            warnings.append(f"{e['company']}: only {n} bullets (expected at least {minimums[i]}).")
    return warnings


def render_keywords(text: str) -> Markup:
    """Escape HTML, then turn <<keyword>> markup into bold spans."""
    if not text:
        return Markup("")
    escaped = html.escape(text)
    escaped = re.sub(r"&lt;&lt;(.+?)&gt;&gt;", r'<strong class="kw">\1</strong>', escaped)
    return Markup(escaped)


app.jinja_env.filters["kw"] = render_keywords


# ---------------------------------------------------------------------------
# Prompt (verbatim from the reference generator; {jobDesc}/{companiesText}/
# {educationText} placeholders are filled in build_prompt()).
# ---------------------------------------------------------------------------

PROMPT_TEMPLATE = r"""
Please make resume with perfect match with "__JOBDESC__".

Target/hiring company to avoid as an experience employer: __TARGETCOMPANY__.
Experience company-name source selected by user: __COMPANYSOURCE__.

CRITICAL — Job title and seniority matching:
- Read the job description carefully. Identify the core job role and seniority level.
- The FIRST line of your output must be: SUBTITLE: <a realistic, clean job title based on the job description>
- IMPORTANT: The SUBTITLE and TITLE fields must be realistic job titles that a real company would use. Strip out any internal level codes (like "L5", "L6", "IC4", "P3"), team/department names (like "- Messaging", "- Platform", "- Growth"), and company-specific jargon. For example:
  * JD says "Data Engineer (L5) - Messaging" -> use "Senior Data Engineer"
  * JD says "Software Engineer II - Payments" -> use "Software Engineer"
  * JD says "Staff SWE, Infrastructure" -> use "Staff Software Engineer"
  * JD says "Frontend Engineer (L4) - Growth" -> use "Frontend Engineer"
  Keep only the standard industry job title that would appear on LinkedIn or a real offer letter.
- The Summary must match the seniority level. If the job says "Junior", do NOT say "9 years of experience as a senior engineer". Instead say something appropriate like "A motivated developer with hands-on experience in..." matching the tone and level of the job description.
- For TITLE in each company: make each job title closely related to the job description role. Use variations that show career progression toward that role. For example if the job is for a Senior Data Engineer, use titles like "Senior Data Engineer", "Data Engineer", "Junior Data Engineer", "Data Analyst" — NOT "Senior Architect" or "Lead Platform Engineer". The most recent company title should be the closest match to the target role. NEVER include level codes, team names, or internal designations in any TITLE field.

Using the style of professional resume (clear headers, concise summary, bullet-pointed achievements with quantified results), generate a complete resume for this candidate.
Format requirements:
- Clear section headers: Summary, Skills, Professional Experience, Education.
- Do NOT include a Certifications section at all.
- Skills: Organize into 4-6 categories that make sense for THIS specific job. Choose category names that logically group the skills required in the job description. Each category label must accurately describe its contents — do NOT force unrelated skills into a mismatched category.
  Rules for skill categories:
  1. Read the job description and identify the major skill areas it requires.
  2. Create category names that genuinely fit those areas (e.g., for a Jira consultant role use "Project Management Tools", "Scripting & Automation", "Data Management" — NOT "Cloud, DevOps & Tools").
  3. Every skill listed under a category must logically belong there. "Microsoft Office Suite" is NOT a Cloud/DevOps tool. "API integrations" is NOT a Cloud/DevOps tool.
  4. Only list actual technologies, tools, platforms, languages, and frameworks — not vague concepts like "Data structures" or "Information Technology".
  5. Include all relevant technologies mentioned in the job description.
  Format: one category per line as "Category Name: skill1, skill2, skill3"
Do NOT use markdown, asterisks, or special characters like ** or ##. Use plain text only.
Do NOT include placeholder lines like [Candidate's Name] or [Phone Number].
Remove all blank lines between skill categories.

Writing style:
- Use clear, professional language with easy-to-read sentences.
- Avoid using asterisks (*) or markdown symbols; use plain text.
- Focus on achievements, impact, and technical depth to attract client's focus.
- SPELLING & GRAMMAR: Your output must be 100% free of spelling errors, grammatical mistakes, and typos. Double-check every word before finalizing. Use correct verb tenses (past tense for previous roles, present tense only for current role). Ensure subject-verb agreement, correct prepositions, and proper punctuation throughout. Common mistakes to avoid: "it's" vs "its", "effect" vs "affect", "insure" vs "ensure", run-on sentences, comma splices, and sentence fragments.
- IMPACT-FIRST writing: Every bullet and every sentence in the Summary must demonstrate measurable impact. Do NOT describe what you were "responsible for" — describe what you ACHIEVED and the result. Recruiters scan for impact numbers within the first 3 words after the verb. Structure: [Verb] [thing] -> [result with number]. Example: "Reduced API response time from 4s to 200ms by refactoring query logic and adding <<Redis>> caching layer, serving 15K daily active users."
- Do NOT use any percentage numbers like "30 percent", "20 percent", "by 40%", "by 25%". Senior engineers describe impact with absolute numbers instead (e.g., "from 12s to 800ms", "processing 3M records daily", "across 40 microservices", "serving 10K concurrent users"). Use counts, durations, throughputs, or scale — never percentages.
- Every bullet MUST mention specific technologies, tools, or frameworks by name.
- Every bullet MUST include a measurable outcome with absolute numbers (not percentages).
- Write like a senior engineer describing real project work, not generic responsibilities.
- IMPORTANT: In the Summary section AND in bullet points, wrap every technical keyword (technology name, framework, tool, protocol, service name, language) inside double angle brackets like <<Kafka>>, <<React>>, <<PostgreSQL>>, <<Kubernetes>>, <<REST API>>, <<CI/CD>>. This is used for formatting only. Example: "Built a streaming pipeline with <<Apache Kafka>> and <<Spark Structured Streaming>>, processing 3M events daily."
- Do NOT use <<>> in the SUBTITLE, COMPANY, DATES, DESCRIPTION, or TITLE fields. Only use <<>> inside the Summary text and the dash (-) bullet point lines.

CRITICAL — Avoid buzzwords and repetition:
- Do NOT use these overused cliché words/phrases ANYWHERE in the resume (Summary, bullets, descriptions — nowhere): "synergy", "leverage", "utilize", "spearheaded", "dynamic", "innovative", "cutting-edge", "best-in-class", "robust", "scalable solutions", "strategic", "proactive", "fast-paced environment", "go-to person", "think outside the box", "team player", "results-driven", "self-starter", "detail-oriented", "passion", "thrive", "adept at", "proven track record", "presentation skills", "strong communication skills", "highly skilled", "extensive experience", "seasoned professional", "committed to", "dedicated to", "responsible for", "in charge of", "various", "numerous", "several projects", "advanced knowledge", "expert in".
- The Summary must contain ZERO vague phrases. Every sentence must state a specific technology, a concrete number, or a measurable domain scope — never generic self-descriptions like "proven track record" or "adept at problem-solving".
- Instead use direct, concrete language. Say exactly what you did and what happened.
- ABSOLUTE RULE — ZERO repeated action verbs: Every single bullet point across the ENTIRE resume must start with a UNIQUE action verb. No verb may appear more than ONCE in the entire document — not even across different companies. Before writing each bullet, check every previous bullet to make sure you have not already used that verb.
- BANNED frequently-overused verbs (do NOT use these AT ALL): "Enhanced", "Improved", "Contributed", "Managed", "Assisted", "Helped", "Worked", "Handled", "Supported", "Utilized", "Developed". These are weak, generic verbs that ATS scanners and resume reviewers penalize.
- Instead, use strong, specific action verbs. Here is a bank of 50+ verbs — pick each one ONLY ONCE:
  Architected, Engineered, Constructed, Formulated, Created, Configured, Established, Assembled, Deployed, Orchestrated, Migrated, Automated, Redesigned, Refactored, Streamlined, Consolidated, Integrated, Pioneered, Implemented, Executed, Transformed, Accelerated, Parallelized, Optimized, Eliminated, Reduced, Resolved, Diagnosed, Overhauled, Modernized, Decoupled, Provisioned, Instrumented, Benchmarked, Standardized, Documented, Authored, Directed, Mentored, Coordinated, Negotiated, Presented, Delivered, Launched, Scaled, Built, Programmed, Coded, Prototyped, Containerized, Virtualized, Partitioned, Indexed, Normalized, Catalogued, Secured, Hardened, Audited, Monitored, Tested, Validated, Certified.
- NEVER repeat the same phrase or sentence structure more than once. Vary sentence length and pattern.
- Do NOT start more than one bullet with the same word across all companies combined.
- In the Summary, avoid generic filler phrases. Every sentence must state a concrete skill, domain, or accomplishment.

CRITICAL — Strong accomplishment-driven bullets with HARD DATA:
- Hiring managers at the senior level look for hard data above everything else. Every bullet MUST contain at least one concrete number. Not vague words like "significantly" or "greatly" — actual numbers.
- Every bullet must follow this formula: [Action verb] + [what you built/did] + [with what technology] + [measurable result with a number].
- Types of hard data to include: record counts (3M rows, 500K events/day), latency (from 12s to 800ms), user counts (serving 10K concurrent users), system scale (across 40 microservices, 15 data sources), time saved (reduced from 3 hours to 10 minutes), team/project scope (led team of 6, across 4 squads), SLA targets (four-nines uptime, p99 latency under 200ms), data volume (processing 2TB daily), infrastructure (deployed to 8-node cluster).
- REALISTIC & CONSISTENT numbers: the figures you invent must be believable for the company's scale and the role's seniority — a junior role at a small firm must NOT claim 5M events per day or a team of 50. Match the magnitude to the seniority (junior: modest scope; senior/staff: large scale and leadership). Keep the numbers within a single bullet consistent with each other, and vary them across bullets — do NOT make every metric a suspiciously round figure (mix values like "1.8M", "240ms", "37 services", "6-person squad").
- BAD (no data): "Managed data pipelines" or "Contributed to system improvements" or "Improved system performance significantly"
- GOOD (hard data): "Architected a real-time ingestion pipeline using <<Apache Kafka>> and <<Spark Structured Streaming>>, processing 5M events daily across 12 partitioned topics with sub-second latency"
- BAD (vague): "Enhanced data quality across the organization"
- GOOD (specific): "Built an automated data validation framework using <<Great Expectations>> and <<Python>>, catching 150+ schema violations per week across 30 production tables"
- Every role must demonstrate IMPACT. For senior roles: show leadership, architecture decisions, scale. For mid roles: show ownership, technical depth, team contributions. For junior roles: show learning speed, hands-on delivery, initiative.
- Each company section should read as a compelling narrative of accomplishments — not a list of generic duties or responsibilities.

ATS (Applicant Tracking System) Optimization — this resume WILL be scored by AI resume scanners. You MUST maximize keyword match score:

STEP 1 — Keyword extraction (do this mentally before writing):
Read the entire job description and extract EVERY keyword, phrase, and term into these categories:
a) Job title variations (e.g., "Software Engineer", "SWE", "Developer")
b) Hard skills & technologies (e.g., "Python", "AWS", "Docker", "Kubernetes")
c) Methodologies & processes (e.g., "Agile", "Scrum", "Kanban", "SDLC", "DevOps", "TDD", "BDD", "CI/CD")
d) Domain/industry terms (e.g., "data pipeline", "ETL", "microservices", "cloud migration", "distributed systems", "machine learning")
e) Certifications & standards mentioned (e.g., "AWS Certified", "ISO 27001", "SOC 2", "PCI DSS", "HIPAA")
f) Soft skill phrases the JD uses (e.g., "cross-functional collaboration", "stakeholder communication", "technical mentorship", "code review")
g) Business context terms (e.g., "SaaS", "B2B", "e-commerce", "fintech", "healthcare", "supply chain")
h) Action/responsibility phrases (e.g., "troubleshoot", "root cause analysis", "capacity planning", "incident response", "performance tuning")

STEP 2 — Keyword placement rules:
- Use a realistic, clean version of the job title (without internal level codes, team names, or company-specific jargon) in the SUBTITLE and the most recent company TITLE.
- EVERY keyword from categories (a) through (h) must appear at least ONCE somewhere in the resume — in Summary, Skills, bullet points, or company descriptions.
- Use the EXACT phrasing from the job description. If the JD says "RESTful APIs", write "RESTful APIs" — not "REST services" or "API development". If the JD says "cross-functional teams", write "cross-functional teams" — not "working with other teams".
- Include both the spelled-out form AND abbreviation where applicable (e.g., "Continuous Integration/Continuous Deployment (CI/CD)", "Amazon Web Services (AWS)", "Software Development Life Cycle (SDLC)").
- Front-load the most important keywords in the Summary (first 2-3 sentences). The Summary should contain at least 60% of the JD's key terms.
- In Skills, list every technology, tool, methodology, framework, and platform mentioned in the job description. Also include closely related industry-standard tools that someone in this role would use, even if not explicitly listed.
- Distribute keywords across ALL sections — do not cluster them only in Skills. Bullet points should naturally embed 2-4 JD keywords each.
- Use standard section headers: Summary, Skills, Professional Experience, Education.

STEP 3 — Industry keyword coverage:
- Beyond what the JD explicitly lists, include standard industry keywords that hiring managers in this field expect. For example, a backend engineer role should mention "API design", "database optimization", "system design", "load balancing", "caching", "logging and monitoring" even if the JD doesn't spell out every one.
- If the JD mentions a cloud provider (AWS/Azure/GCP), include 3-5 specific services from that provider in your bullets (e.g., for AWS: "EC2", "S3", "Lambda", "RDS", "CloudWatch").
- If the JD mentions "Agile" or "Scrum", reference sprint planning, retrospectives, or iterative development in at least one bullet.
- If the JD mentions collaboration or leadership, include bullets about code reviews, mentoring junior developers, or leading technical discussions.

Summary:
Write 4-6 impactful sentences. Match the tone to the job description seniority:
- If the job is senior/staff level: mention years of experience (use the number __YOE__) and deep expertise.
- If the job is mid-level: mention solid experience and growing expertise.
- If the job is junior/entry level: focus on enthusiasm, hands-on skills, and foundational knowledge — do NOT mention "9 years of experience" for a junior role.
Always mention the specific tech skills required in the job description.
Wrap technical keywords in <<>> in the Summary too, e.g., "Experienced in <<Python>>, <<Jira>>, and <<REST API>> integrations."

Skills:

Professional Experience:
For each company, use this EXACT structured format (one field per line):
COMPANY: Company Name, Remote
DATES: Month YYYY – Month YYYY (e.g., "February 2022 – January 2024", "March 2019 – May 2022". Always use full month name, NOT numeric format like "2024-02")
CRITICAL — DATES must be taken VERBATIM from the company list provided below. Use the exact From and To dates as given. Do NOT invent, round, or change any date. If the To value is "Present" or empty, write "Present".
DESCRIPTION: One-sentence company description that highlights work relevant to the job description. Frame the company's work using keywords and domain language from the job posting. For example, if the job is about "data engineering", describe the company as "A technology firm specializing in enterprise data platform solutions and cloud-based analytics infrastructure" — not just "A software company". Make it sound like every company the candidate worked at was doing work closely related to the target job.
TITLE: Detailed Job Title
- bullet point 1
- bullet point 2
...
IMPORTANT: For the COMPANY field, always use the company name followed by "Remote" — do NOT use City or State. Example: "COMPANY: Cibirix, Remote"

MANDATORY bullet count per role — you MUST meet these minimums or the output is invalid:
- Most recent company (latest dates): MINIMUM 10 bullet points, ideally 12. Do NOT write fewer than 10.
- Second most recent company: MINIMUM 8 bullet points, ideally 10. Do NOT write fewer than 8.
- Third company: MINIMUM 6 bullet points. Do NOT write fewer than 6.
- Fourth and older companies: MINIMUM 4 bullet points.
Count your bullets for each company before moving to the next. If a company has fewer than the minimum, add more bullets before proceeding.
This is the most important formatting rule — recruiters judge experience depth by bullet count.

Write bullet points for each role. Each bullet MUST:
1. Start with a strong action verb (Architected, Engineered, Optimized, Orchestrated, Migrated, Automated, Implemented, Redesigned, Built, Constructed, Configured, Established, Formulated, Consolidated, Refactored, Streamlined, Integrated, Deployed).
2. Name specific technologies wrapped in <<>> (e.g., "using <<Kafka>>, <<Spark>>, and <<Airflow>>").
3. Describe the technical problem or project scope concretely (e.g., "real-time event streaming pipeline processing 5M events per day" not just "data pipeline").
4. End with a measurable result using absolute numbers — NEVER percentages. Use counts, durations, throughput, or scale (e.g., "reducing query time from 12s to 800ms", "serving 10K concurrent users", "across 40 microservices").
Do NOT write vague bullets like "Improved system performance" — always specify what system, what technology, and what measurable result.
ABSOLUTE RULE: The resume must contain ZERO percentages. No "40%", no "25 percent", no "by 50%", no "increase of 30%". If you catch yourself writing a percentage, replace it with an absolute number (e.g., instead of "improved speed by 40%" write "improved speed from 5s to 800ms"). Scan your entire output and remove every single percentage before finishing.

Example of GOOD bullets:
- Architected a real-time data ingestion pipeline using <<Apache Kafka>> and <<Spark Structured Streaming>>, processing 3M events daily with sub-second latency across 12 partitioned topics.
- Migrated 15 legacy monolithic <<ETL>> workflows to <<Apache Airflow>> DAGs on <<AWS MWAA>>, cutting pipeline recovery time from 3 hours to under 10 minutes.
- Built a medallion architecture (Bronze/Silver/Gold) in <<Microsoft Fabric>> Lakehouse, enabling self-service analytics for 50 business users and reducing report generation from 4 hours to 15 minutes.

OVERRIDE — use this company-name rule instead of any earlier company-name rule:
- For OpenAI company-name source, choose a different employer company name for each experience slot below.
- The company names must fit the job description's industry/domain.
- Choose believable mid-sized companies: not famous, not Fortune 500, not Big Tech, not tiny startups, and not household names.
- Do NOT use the hiring company or target company from the job description.
- Do NOT use the target/hiring company listed above.
- Do NOT use any company name that appears in the job description.
- Do NOT use generic experience slot labels as final company names.
- If the selected company-name source is OpenAI, do NOT use the profile company names as final COMPANY names; generate replacement employer names.
- If the selected company-name source is Profile, keep the profile company names in the COMPANY fields.
- Company names should sound like real mid-market firms in the field, but should be understated and not attention-grabbing.
- Use the exact number of experience slots below. Keep the dates and order from those slots.
- Output exactly the company/experience count shown below. Do not add, skip, or merge experience blocks.

__COMPANIESTEXT__

Please show experience sorted by latest company.

Education:
For each degree, use this EXACT structured format (one field per line):
UNIVERSITY: University Name, City, State
GRADUATION: Month YYYY
DEGREE: Full degree name (e.g., Bachelor of Science in Computer Science)
COURSEWORK: course1 • course2 • course3

__EDUCATIONTEXT__
"""


def _parse_ym(value: str):
    value = (value or "").strip()
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", value)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.match(r"^(\d{4})$", value)
    if m:
        return int(m.group(1)), 1
    return None


def years_of_experience(profile: dict) -> int:
    """Total career span (earliest start -> latest end) from the real company
    dates, rounded to the nearest year. Replaces the prompt's hardcoded '9'."""
    today = date.today()
    starts, ends = [], []
    for c in profile.get("companies", []):
        s = _parse_ym(c.get("start", ""))
        if s:
            starts.append(s)
        end_raw = (c.get("end", "") or "").strip()
        if end_raw.lower() in ("", "present"):
            ends.append((today.year, today.month))
        else:
            ends.append(_parse_ym(end_raw) or (today.year, today.month))
    if not starts:
        return 9
    sy, sm = min(starts)
    ey, em = max(ends)
    months = (ey - sy) * 12 + (em - sm)
    years = months // 12 + (1 if months % 12 >= 6 else 0)
    return max(1, years)


def build_companies_text(profile: dict) -> str:
    lines = []
    lines.append(f"Total experience/company count: {len(profile['companies'])}.")
    for i, c in enumerate(profile["companies"], start=1):
        lines.append(
            f"- Experience slot {i}: "
            f"Profile company name: {c.get('name', '')}. "
            f"Location to keep: {c.get('location', '')}. "
            f"From: {fmt_date(c.get('start',''))}, "
            f"To: {fmt_date(c.get('end',''), is_end=True)}"
        )
    return "\n".join(lines)


def build_education_text(profile: dict) -> str:
    lines = []
    for e in profile["education"]:
        lines.append(
            f"- {e['university']}, {e.get('location', '')}, "
            f"{fmt_date(e.get('start',''))} - {fmt_date(e.get('end',''), is_end=True)}. "
            f"Choose a DEGREE that best fits the job description."
        )
    return "\n".join(lines)


def build_prompt(profile: dict, job_description: str, target_company: str = "",
                 company_source: str = "profile") -> str:
    source_label = "OpenAI" if company_source == "openai" else "Profile"
    return (PROMPT_TEMPLATE
            .replace("__JOBDESC__", job_description)
            .replace("__TARGETCOMPANY__", target_company or "Not provided; infer it from the job description if present")
            .replace("__COMPANYSOURCE__", source_label)
            .replace("__YOE__", str(years_of_experience(profile)))
            .replace("__COMPANIESTEXT__", build_companies_text(profile))
            .replace("__EDUCATIONTEXT__", build_education_text(profile)))


# ---------------------------------------------------------------------------
# Parse the model's plain-text resume into structured data
# ---------------------------------------------------------------------------

_HEADERS = {"summary": "summary", "skills": "skills",
            "professional experience": "experience", "experience": "experience",
            "education": "education"}


def _field(line: str) -> str:
    return line.split(":", 1)[1].strip() if ":" in line else ""


def _strip_kw(text: str) -> str:
    """Remove <<>> markup. Used for fields that should render as plain text
    (skills, titles, descriptions) — the markup belongs only in summary/bullets."""
    return (text or "").replace("<<", "").replace(">>", "")


def parse_resume_text(text: str) -> dict:
    subtitle = ""
    summary_parts, skills, experience, education = [], [], [], []
    section = None
    cur_exp = cur_edu = None

    def flush():
        nonlocal cur_exp, cur_edu
        if cur_exp:
            experience.append(cur_exp); cur_exp = None
        if cur_edu:
            education.append(cur_edu); cur_edu = None

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        if line.upper().startswith("SUBTITLE:"):
            subtitle = _field(line)
            continue

        # A section header is a line that is JUST the header word, optionally
        # followed by a colon and nothing else (e.g. "Summary" or "Skills:").
        parts = line.split(":", 1)
        head = parts[0].strip().lower()
        after = parts[1].strip() if len(parts) > 1 else ""
        if head in _HEADERS and not after:
            flush()
            section = _HEADERS[head]
            continue

        if section == "summary":
            summary_parts.append(line)

        elif section == "skills":
            if ":" in line:
                label, items = line.split(":", 1)
                vals = [_strip_kw(s).strip() for s in items.split(",") if s.strip()]
                if vals:
                    skills.append({"label": _strip_kw(label).strip(), "skills": vals})

        elif section == "experience":
            U = line.upper()
            if U.startswith("COMPANY:"):
                flush()
                cur_exp = {"company": _field(line), "dates": "", "description": "",
                           "title": "", "bullets": []}
            elif cur_exp is None:
                continue
            elif U.startswith("DATES:"):
                cur_exp["dates"] = _field(line)
            elif U.startswith("DESCRIPTION:"):
                cur_exp["description"] = _strip_kw(_field(line))
            elif U.startswith("TITLE:"):
                cur_exp["title"] = _strip_kw(_field(line))
            elif line[0] in "-•*":
                cur_exp["bullets"].append(line[1:].strip())

        elif section == "education":
            U = line.upper()
            if U.startswith("UNIVERSITY:"):
                flush()
                cur_edu = {"university": _field(line), "graduation": "",
                           "degree": "", "coursework": ""}
            elif cur_edu is None:
                continue
            elif U.startswith("GRADUATION:"):
                cur_edu["graduation"] = _field(line)
            elif U.startswith("DEGREE:"):
                cur_edu["degree"] = _strip_kw(_field(line))
            elif U.startswith("COURSEWORK:"):
                cur_edu["coursework"] = _strip_kw(_field(line))

    flush()
    return {"subtitle": subtitle, "summary": " ".join(summary_parts),
            "skills": skills, "experience": experience, "education": education}


def _norm_company(name: str) -> str:
    """Normalize a company name for matching: take the part before the first
    comma (drops the ', Remote' the model appends) and keep only alphanumerics.
    'BYLTAX, Remote' and 'byltax' both become 'byltax'."""
    head = (name or "").split(",", 1)[0]
    return re.sub(r"[^a-z0-9]+", "", head.lower())


def clean_company_name(name: str) -> str:
    """Use only the company name part from model output like 'Acme, Remote'."""
    return (name or "").split(",", 1)[0].strip()


def invalid_openai_company_names(parsed_exp: list, profile: dict,
                                 target_company: str = "") -> list:
    """Problems that make OpenAI-generated company names unsafe to use."""
    problems = []
    profile_names = {_norm_company(c.get("name", "")) for c in profile["companies"]}
    profile_names.discard("")
    target = _norm_company(target_company)
    seen = set()
    if len(parsed_exp) < len(profile["companies"]):
        problems.append(
            f"returned {len(parsed_exp)} company blocks for "
            f"{len(profile['companies'])} profile companies"
        )
    for i, c in enumerate(profile["companies"]):
        ai = parsed_exp[i] if i < len(parsed_exp) else {}
        company = clean_company_name(ai.get("company", ""))
        norm = _norm_company(company)
        if not norm:
            problems.append(f"slot {i + 1} has no generated company name")
        elif norm in profile_names:
            problems.append(f"slot {i + 1} reused profile company '{c.get('name', '')}'")
        elif target and norm == target:
            problems.append(f"slot {i + 1} reused target company '{target_company}'")
        elif norm in seen:
            problems.append(f"slot {i + 1} duplicated generated company '{company}'")
        seen.add(norm)
    return problems


def tailor_resume(profile: dict, job_description: str, target_company: str = "",
                  company_source: str = "profile") -> dict:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or api_key.startswith("sk-paste"):
        raise RuntimeError("No OpenAI API key configured. Edit .env and add your key.")
    client = OpenAI(api_key=api_key)

    resp = client.chat.completions.create(
        model=MODEL,
        temperature=0.6,
        messages=[
            {"role": "system", "content":
             "You are an expert resume writer. Follow the user's instructions exactly "
             "and output plain text only, in the precise structure requested."},
            {"role": "user", "content": build_prompt(profile, job_description, target_company, company_source)},
        ],
    )
    parsed = parse_resume_text(resp.choices[0].message.content or "")
    if company_source == "openai":
        problems = invalid_openai_company_names(
            parsed["experience"], profile, target_company)
        if problems:
            retry_prompt = (
                build_prompt(profile, job_description, target_company, company_source)
                + "\n\nSTRICT CORRECTION REQUIRED:\n"
                + "The previous company names were invalid because: "
                + "; ".join(problems)
                + ". Generate the full resume again. Use exactly "
                + str(len(profile["companies"]))
                + " experience/company blocks. Do not use profile company names, "
                + "generic slot labels, duplicate company names, or the target/hiring company."
            )
            resp = client.chat.completions.create(
                model=MODEL,
                temperature=0.6,
                messages=[
                    {"role": "system", "content":
                     "You are an expert resume writer. Follow the user's instructions exactly "
                     "and output plain text only, in the precise structure requested."},
                    {"role": "user", "content": retry_prompt},
                ],
            )
            parsed = parse_resume_text(resp.choices[0].message.content or "")
            problems = invalid_openai_company_names(
                parsed["experience"], profile, target_company)
            if problems:
                raise RuntimeError(
                    "OpenAI company-name generation failed: " + "; ".join(problems)
                )

    # Overlay AI prose onto the FIXED facts (company/dates/location/school verbatim).
    # Match each profile company to the AI block with the SAME name rather than
    # trusting list position — if the model reorders/drops a company, position
    # matching would attach bullets to the wrong employer. Fall back to position
    # only when no name matches (and that positional block isn't already taken).
    parsed_exp = parsed["experience"]
    experience = []
    for i, c in enumerate(profile["companies"]):
        ai = parsed_exp[i] if i < len(parsed_exp) else {}
        ai_company = clean_company_name(ai.get("company", ""))
        company = ai_company if company_source == "openai" and ai_company else c["name"]
        experience.append({
            "company": company,
            "location": c.get("location", ""),
            "dates": fmt_range(c.get("start", ""), c.get("end", "")),
            "title": ai.get("title", ""),
            "description": ai.get("description", ""),
            "bullets": [strip_percentages(b) for b in ai.get("bullets", [])],
        })

    education = []
    for i, e in enumerate(profile["education"]):
        ai = parsed["education"][i] if i < len(parsed["education"]) else {}
        education.append({
            "university": e["university"],
            "location": e.get("location", ""),
            "dates": fmt_range(e.get("start", ""), e.get("end", "")),
            "degree": ai.get("degree", ""),
            "coursework": ai.get("coursework", ""),
        })

    return {"subtitle": parsed["subtitle"],
            "summary": strip_percentages(parsed["summary"]),
            "skills": parsed["skills"], "experience": experience, "education": education}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    active = get_active_id()
    company_source = session.get("company_source", "profile")
    if company_source not in ("profile", "openai"):
        company_source = "profile"
    return render_template("index.html", profiles=list_profile_names(list_profiles()),
                           active=active, profile=load_profile(active),
                           company_source=company_source)


@app.route("/profile", methods=["GET", "POST"])
def profile():
    if request.method == "POST":
        pid = request.form.get("profile_id", "").strip() or get_active_id()
        if not pid:  # no profile to save into — create one from the form name
            pid = create_profile(request.form.get("name", "").strip() or "profile")
        save_profile(parse_profile_form(request.form), pid)
        set_active_id(pid)
        flash("Profile saved.", "success")
        return redirect(url_for("profile", id=pid))

    profiles = list_profiles()
    pid = request.args.get("id", "").strip() or get_active_id()
    if pid:
        set_active_id(pid)
    profile_data = load_profile(pid)
    names = list_profile_names(profiles)
    return render_template("profile.html", profiles=profiles, current=pid,
                           profile=profile_data, names=names)


@app.route("/profile/new", methods=["POST"])
def profile_new():
    pid = create_profile(request.form.get("new_name", "").strip() or "New profile")
    set_active_id(pid)
    flash("Profile created — fill in the details and save.", "success")
    return redirect(url_for("profile", id=pid))


@app.route("/profile/delete", methods=["POST"])
def profile_delete():
    pid = request.form.get("profile_id", "").strip()
    if pid:
        delete_profile(pid)
        flash("Profile deleted.", "success")
    remaining = list_profiles()
    set_active_id(remaining[0]["id"] if remaining else None)
    return redirect(url_for("profile"))


def render_pdf(profile_data: dict, tailored: dict) -> bytes:
    html_doc = render_template("resume.html", profile=profile_data, tailored=tailored,
                               preview=False)
    return HTML(string=html_doc, base_url=str(BASE_DIR)).write_pdf()


def _slug(text: str) -> str:
    """Filesystem-safe filename segment: alphanumerics, runs collapsed to '_'."""
    return re.sub(r"[^A-Za-z0-9]+", "_", text or "").strip("_")


def build_filename(profile_data: dict, company: str) -> str:
    """e.g. 'Spencer_Aguas_Acme_1735649895.pdf', or
    'Spencer_Aguas_1735649895.pdf' when no company is given. The Unix timestamp
    keeps each download unique; if two resumes land in the same second, a numeric
    suffix is added so a PDF is never silently overwritten."""
    name = _slug(profile_data.get("name", "")) or "resume"
    stamp = int(time.time())
    company = _slug(company)
    base = "_".join([name] + ([company] if company else []) + [str(stamp)])
    candidate = base + ".pdf"
    n = 2
    while (RESUMES_DIR / candidate).exists():
        candidate = f"{base}_{n}.pdf"
        n += 1
    return candidate


def jd_filename(pdf_filename: str) -> str:
    """The job-description text file that pairs with a resume PDF: same base name,
    '.txt' extension (e.g. 'Spencer_Aguas_Acme_123.pdf' -> '..._123.txt')."""
    return Path(pdf_filename).with_suffix(".txt").name


def save_job_description(pdf_filename: str, job_description: str, company: str,
                         job_title: str, job_link: str) -> None:
    """Save the job description next to its resume PDF as a self-documenting .txt,
    so you can always see what each resume was tailored to. Best-effort: a write
    failure here must never lose the generated PDF."""
    try:
        header = [
            f"Company:   {company or '—'}",
            f"Job title: {job_title or '—'}",
            f"Job link:  {job_link or '—'}",
            f"Saved:     {time.strftime('%Y-%m-%d %H:%M')}",
            f"Resume:    {pdf_filename}",
            "=" * 60,
            "",
        ]
        text = "\n".join(header) + job_description.strip() + "\n"
        (RESUMES_DIR / jd_filename(pdf_filename)).write_text(text, encoding="utf-8")
    except OSError as exc:
        app.logger.warning("Could not save job description: %s", exc)


def _save_workbook(wb, path: Path) -> None:
    """Save atomically: write to a unique temp file in the same folder, then
    os.replace it into place. A crash or concurrent write can never leave a
    half-written / truncated workbook.

    The temp name is unique per save (mkstemp) so two requests writing the same
    profile's log at once don't clobber a shared temp file, and a stale temp from
    a past crash can never block future saves.

    On Windows the final replace can hit a lock on the destination and raise
    PermissionError — WinError 5 (access denied, e.g. OneDrive/antivirus) or
    WinError 32 (sharing violation, the .xlsx is open in Excel). Transient locks
    clear in milliseconds, so retry with a short backoff. A persistent lock still
    raises, with a clear message telling the user to close the file."""
    fd, tmp_name = tempfile.mkstemp(
        dir=str(path.parent), prefix=path.stem + "_", suffix=".tmp.xlsx")
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        last_err = None
        for attempt in range(6):
            try:
                os.replace(tmp, path)
                return  # success — tmp no longer exists, finally is a no-op
            except PermissionError as e:  # WinError 5 / 32 on a locked destination
                last_err = e
                time.sleep(0.25 * (attempt + 1))
        raise PermissionError(
            f"Could not save {path.name}: the file is locked. Close it in Excel "
            f"(and consider moving the project out of OneDrive) and try again."
        ) from last_err
    finally:
        if tmp.exists():  # only on failure; clean up the orphaned temp file
            try:
                os.remove(tmp)
            except OSError:
                pass


def log_resume(pid: str, job_title: str, url: str, company: str, filename: str,
               progress: str = "", note: str = "") -> None:
    """Append a row to the profile's log (resumes_log_<pid>.xlsx). Columns: Date,
    Job Title, URL, Company, Filename, Progress, Note. Never breaks the download."""
    path = resumes_log_path(pid)
    try:
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
            if ws.max_row == 0 or [c.value for c in ws[1]] != LOG_HEADERS:
                ws.insert_rows(1)
                for col, header in enumerate(LOG_HEADERS, start=1):
                    ws.cell(row=1, column=col, value=header)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resumes Log"
            ws.append(LOG_HEADERS)
        ws.append([time.strftime("%Y-%m-%d %H:%M:%S"), job_title, url, company,
                   filename, progress, note])
        _save_workbook(wb, path)
    except Exception as exc:  # noqa: BLE001 — never block the download on logging
        app.logger.warning("Could not write resumes log: %s", exc)


def resume_sort_key(date_value: str, filename: str, row: int) -> tuple:
    """Newest-first sort key. Prefer the Unix timestamp embedded in filenames."""
    m = re.search(r"_(\d{10})(?:_\d+)?\.pdf$", filename or "")
    if m:
        return (int(m.group(1)), row)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return (int(time.mktime(time.strptime(str(date_value), fmt))), row)
        except (TypeError, ValueError):
            pass
    return (0, row)


def read_log_rows(pid: str) -> list:
    """A profile's logged resumes as dicts, plus whether each PDF still exists."""
    if not pid:
        return []
    path = resumes_log_path(pid)
    if not path.exists():
        return []
    # read_only=True keeps the file handle open until wb.close(); on Windows a
    # leaked handle locks the file and the next write fails (WinError 32). Always
    # close it in a finally so the workbook is never left open.
    wb = None
    rows = []
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not r or all(c is None for c in r):
                continue
            r = list(r) + [None] * (7 - len(r))
            filename = str(r[4] or "")
            progress = str(r[5] or "").strip()
            date_value = str(r[0] or "")
            rows.append({
                "row": idx,  # actual worksheet row, used to update Progress in place
                "profile_id": pid,
                "sort_key": resume_sort_key(date_value, filename, idx),
                "date": date_value,
                "job_title": str(r[1] or ""),
                "url": str(r[2] or ""),
                "company": str(r[3] or ""),
                "filename": filename,
                "progress": progress,
                # Match any case/variant ("REJECTED", "Rejected (phone only)", ...)
                "is_rejected": progress.upper().startswith("REJECT"),
                "note": str(r[6] or ""),
                "exists": bool(filename) and (RESUMES_DIR / filename).exists(),
                # The paired job-description text file, if it was saved alongside.
                "jd": jd_filename(filename) if filename else "",
                "jd_exists": bool(filename)
                and (RESUMES_DIR / jd_filename(filename)).exists(),
            })
    except Exception as exc:  # noqa: BLE001
        app.logger.warning("Could not read resumes log: %s", exc)
        return []
    finally:
        if wb is not None:
            wb.close()
    return rows


@app.route("/generate", methods=["POST"])
def generate():
    job_description = request.form.get("job_description", "").strip()
    if not job_description:
        flash("Please paste a job description first.", "error")
        return redirect(url_for("index"))

    pid = request.form.get("profile_id", "").strip() or get_active_id()
    profile_data = load_profile(pid)
    if pid:
        set_active_id(pid)  # remember the last profile used
    if not profile_data["companies"]:
        flash("Add your experience slots on the Profile page first.", "error")
        return redirect(url_for("profile"))

    company = request.form.get("company_name", "").strip()
    job_link = request.form.get("job_link", "").strip()
    job_title = request.form.get("job_title", "").strip()
    company_source = request.form.get("company_source", "profile").strip()
    if company_source not in ("profile", "openai"):
        company_source = "profile"
    session["company_source"] = company_source

    try:
        tailored = tailor_resume(profile_data, job_description, company, company_source)
        for w in build_warnings(tailored):
            app.logger.warning("Resume parse: %s", w)
        pdf_bytes = render_pdf(profile_data, tailored)
    except Exception as exc:  # noqa: BLE001
        flash(f"Generation failed: {exc}", "error")
        return redirect(url_for("index"))

    # Save the finished PDF into the project's resumes/ folder (not the browser
    # Downloads folder), then return to the page with a confirmation.
    filename = build_filename(profile_data, company)

    RESUMES_DIR.mkdir(exist_ok=True)
    (RESUMES_DIR / filename).write_bytes(pdf_bytes)
    save_job_description(filename, job_description, company, job_title, job_link)
    log_resume(pid, job_title, job_link, company, filename)

    flash(f"Saved to resumes/{filename}", "success")
    return redirect(url_for("index"))


def _resumes_redirect(pid: str, query: str):
    args = {}
    if query:
        args["q"] = query
    if pid:
        args["profile"] = pid
    return redirect(url_for("resumes", **args))


@app.route("/resumes")
def resumes():
    _migrate_legacy_log()
    profiles = list_profiles()
    pid = request.args.get("profile", "").strip() or get_active_id()
    if pid:
        set_active_id(pid)
    query = request.args.get("q", "").strip()
    current_profile = load_profile(pid)
    current_name = current_profile.get("name", "")
    names = list_profile_names(profiles)
    selected_profiles = profiles_for_name(current_name, profiles)
    rows = []
    for p in selected_profiles:
        for row in read_log_rows(p["id"]):
            row["profile_id"] = p["id"]
            rows.append(row)
    rows.sort(key=lambda r: r["sort_key"], reverse=True)
    total = len(rows)
    if query:
        q = query.lower()
        rows = [r for r in rows
                if q in r["job_title"].lower()
                or q in r["company"].lower()
                or q in r["url"].lower()]
    return render_template("resumes.html", rows=rows, q=query, total=total,
                           shown=len(rows), profiles=profiles, current=pid,
                           names=names, current_name=current_name)


PROGRESS_VALUES = ("", "Rejected")


@app.route("/resumes/progress", methods=["POST"])
def update_progress():
    """Set a row's Progress to '' or 'Rejected' (column 6) in the profile's log."""
    row = request.form.get("row", type=int)
    progress = request.form.get("progress", "").strip()
    query = request.form.get("q", "").strip()
    pid = request.form.get("profile_id", "").strip() or get_active_id()
    if progress not in PROGRESS_VALUES:
        progress = ""
    path = resumes_log_path(pid) if pid else None
    if row and row >= 2 and path and path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            if row <= ws.max_row:
                # Assign .value directly: ws.cell(..., value=None) is a no-op in
                # openpyxl (it won't clear an existing value), so blanking fails.
                ws.cell(row=row, column=6).value = progress or None
                _save_workbook(wb, path)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Could not update progress: %s", exc)
            flash("Could not update progress (is the Excel file open elsewhere?).", "error")
    return _resumes_redirect(pid, query)


@app.route("/resumes/note", methods=["POST"])
def update_note():
    """Set a row's free-text Note (column 7) in the profile's log."""
    row = request.form.get("row", type=int)
    note = request.form.get("note", "").strip()
    query = request.form.get("q", "").strip()
    pid = request.form.get("profile_id", "").strip() or get_active_id()
    path = resumes_log_path(pid) if pid else None
    if row and row >= 2 and path and path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            if row <= ws.max_row:
                ws.cell(row=row, column=7).value = note or None
                _save_workbook(wb, path)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Could not update note: %s", exc)
            flash("Could not update note (is the Excel file open elsewhere?).", "error")
    return _resumes_redirect(pid, query)


@app.route("/resumes/delete", methods=["POST"])
def delete_resume():
    """Delete a log row (and its PDF if saved in resumes/) from the profile's log."""
    row = request.form.get("row", type=int)
    query = request.form.get("q", "").strip()
    pid = request.form.get("profile_id", "").strip() or get_active_id()
    path = resumes_log_path(pid) if pid else None
    if row and row >= 2 and path and path.exists():
        try:
            wb = load_workbook(path)
            ws = wb.active
            if row <= ws.max_row:
                filename = ws.cell(row=row, column=5).value
                ws.delete_rows(row, 1)
                _save_workbook(wb, path)
                # Remove the saved PDF and its companion job-description .txt too
                # (Path(...).name blocks path traversal).
                if filename:
                    safe = Path(str(filename)).name
                    for f in (RESUMES_DIR / safe, RESUMES_DIR / jd_filename(safe)):
                        if f.exists():
                            f.unlink()
                flash("Deleted resume entry.", "success")
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Could not delete row: %s", exc)
            flash("Could not delete (is the Excel file open elsewhere?).", "error")
    return _resumes_redirect(pid, query)


@app.route("/resumes/file/<path:filename>")
def resume_file(filename):
    """Serve a saved PDF from resumes/. send_from_directory blocks path traversal."""
    if not (RESUMES_DIR / filename).exists():
        abort(404)
    return send_from_directory(RESUMES_DIR, filename)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
